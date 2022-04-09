# -*- coding: utf-8 -*-
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.font as tf
import tkinter.messagebox as mb
import openpyxl as xl
import sys
from tkinter.filedialog import askopenfilename
import pymysql
import datetime as dt
import pyautogui as pg
from tkinter.ttk import *
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
import os
import qrcode as qr
from PIL import Image

"""
십일조 주일헌금, 주정헌금 감사헌금 선교헌금 건축헌금 특별헌금 종합
"""


def church():
    s = smtplib.SMTP("smtp.gmail.com", 587)
    s.starttls()
    f = open("info.txt", 'r', encoding='utf-8')
    data = f.readlines()
    유저 = data[1].rstrip('\n')
    비밀번호 = data[3].rstrip('\n')
    데이터베이스 = data[5].rstrip('\n')
    DB테이블 = data[7].rstrip('\n')
    s.login(data[9].rstrip('\n'), data[11].rstrip('\n'))
    보내는메일 = data[9].rstrip('\n')
    받는메일 = data[13].rstrip('\n')
    누계테이블 = data[15].rstrip('\n')
    출석테이블 = data[17].rstrip('\n')
    f.close()
    church_db = pymysql.connect(host="localhost", user=유저, password=비밀번호, db=데이터베이스, charset="utf8")
    cursor = church_db.cursor(pymysql.cursors.DictCursor)
    목장선택 = (
        "선택해주세요", "목사님가정", "전도사님가정", "청년목장", "코아목장", "베데스다", "샬롬목장", "예닮목장", "사랑목장", "우리목장", "푸른초장", "할렐루야", "섬김목장",
        "없음")
    직분선택 = ("선택해주세요", "목회자", "장로", "권사", "집사", "성도", "청년", "학생", "아동", "없음")
    년선택, 월선택, 일선택 = (), (), ()
    # noinspection PyUnusedLocal
    교인코드길이 = 22
    헌금형식 = {"1": "십일조", "2": "주일헌금", "3": "건축헌금", "4": "선교헌금", "5": "감사헌금", "6": "특별헌금"}
    # noinspection PyUnusedLocal
    for i in range(int(dt.datetime.today().year), 1899, -1):
        년선택 += (str(i),)
    for i in range(1, 13):
        월선택 += (str(i).zfill(2),)
    for i in range(1, 32):
        일선택 += (str(i).zfill(2),)

    def Main():

        def 자동입력():
            try:
                info_file = xl.load_workbook(askopenfilename(filetypes=[("Excel files", "*.xlsx")]))
                excel_cursor = info_file.active
                겉테이블 = []
                for c in range(2, excel_cursor.max_row + 1):
                    교인코드 = excel_cursor['A' + str(c)].value.strftime("%Y%m%d%H%M%S") + excel_cursor[
                        'H' + str(c)].value.strftime("%Y%m%d")
                    sql = "SELECT * FROM " + DB테이블 + " WHERE 교인코드 = '" + 교인코드 + "';"
                    code = cursor.execute(sql)
                    if code == 0:
                        미니테이블 = [교인코드]

                        이름 = excel_cursor['C' + str(c)].value.replace(" ", "")
                        미니테이블.append(이름)

                        휴대전화번호 = excel_cursor['D' + str(c)].value
                        if 휴대전화번호 != "" and 휴대전화번호 is not None:
                            filter(str.isdigit, 휴대전화번호)
                        else:
                            휴대전화번호 = "없음"
                        미니테이블.append(휴대전화번호)

                        목장 = excel_cursor['F' + str(c)].value
                        if excel_cursor['G' + str(c)].value == "청년":
                            미니테이블.append("청년목장")
                        elif excel_cursor['G' + str(c)].value == "목회자":
                            미니테이블.append("목사님가정")
                        elif excel_cursor['G' + str(c)].value == "권사":
                            미니테이블.append("전도사님가정")
                        else:
                            미니테이블.append(목장)

                        직분 = excel_cursor['G' + str(c)].value
                        미니테이블.append(직분)

                        생년월일 = excel_cursor['H' + str(c)].value.strftime("%F")
                        미니테이블.append(생년월일)

                        가족 = excel_cursor['I' + str(c)].value
                        if 가족 == "" or 가족 is None:
                            가족 = "없음"
                        미니테이블.append(가족)

                        주소 = excel_cursor['J' + str(c)].value
                        미니테이블.append(주소)

                        목장지기여부 = excel_cursor['K' + str(c)].value
                        if 목장지기여부 == "맞음":
                            목장지기여부 = 'Y'
                        else:
                            목장지기여부 = 'N'
                        미니테이블.append(목장지기여부)

                        집전화번호 = str(excel_cursor['E' + str(c)].value)
                        if 집전화번호 == "None":
                            집전화번호 = "없음"
                        else:
                            filter(str.isdigit, 집전화번호)
                        미니테이블.append(집전화번호)

                        침례세례일 = excel_cursor['L' + str(c)].value
                        침례세례교회 = excel_cursor['M' + str(c)].value
                        침례세례목사 = excel_cursor['N' + str(c)].value
                        if 침례세례일 is not None:
                            침례세례일 = 침례세례일.strftime("%F")
                        else:
                            침례세례일 = "없음"
                        미니테이블.append(침례세례일)

                        if 침례세례교회 is None:
                            침례세례교회 = "없음"
                        미니테이블.append(침례세례교회)

                        if 침례세례목사 is None:
                            침례세례목사 = "없음"
                        미니테이블.append(침례세례목사)

                        미니테이블.append('Y')

                        겉테이블.append(미니테이블)
                    else:
                        print("이미 있음")
                sql = "INSERT INTO `" + DB테이블 + "`(교인코드, 이름, 휴대전화번호, 목장, 직분, 생년월일, 가족, 주소, 목장지기여부, 집전화번호, 침례세례일, 침례세례받은교회, 침례세례목사, 개인정보활용동의) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);"
                cursor.executemany(sql, 겉테이블)
                church_db.commit()
                mb.showinfo("완료", "자동적으로 입력되었습니다.")
            except:
                mb.showerror("오류", "자동 입력 중 오류가 발생했습니다.")

        def 수동입력():
            주석폰트 = tf.Font(family="맑은 고딕", size=8)
            라벨폰트 = tf.Font(family="맑은 고딕", size=12)
            버튼폰트 = tf.Font(family="맑은 고딕", size=10)
            테이블 = {"교인코드": dt.datetime.today().strftime("%Y%m%d%H%M%S"),
                   "이름": "", "휴대전화번호": "", "목장": "", "직분": "",
                   "생년월일": "", "가족": "", "주소": "", "목장지기여부": "",
                   "집전화번호": "", "침례세례일": "", "침례세례받은교회": "",
                   "침례세례목사": "", "개인정보활용동의": ""}

            def 저장():
                if 동의버튼['state'] == tk.DISABLED:
                    테이블["개인정보활용동의"] = "Y"
                    if 이름입력.get() != '':
                        테이블["이름"] = 이름입력.get().replace(" ", "")
                        if 휴대전화번호입력.get() != '' and len(휴대전화번호입력.get()) >= 11:
                            테이블["휴대전화번호"] = 휴대전화번호입력.get()[:11]
                            filter(str.isdigit, 테이블["휴대전화번호"])
                        else:
                            테이블["휴대전화번호"] = "없음"
                        if 집전화번호입력.get() != '':
                            테이블["집전화번호"] = 집전화번호입력.get()
                        else:
                            테이블["집전화번호"] = "없음"
                        if 직분콤보박스.get() != "선택해주세요":
                            테이블["직분"] = 직분콤보박스.get()
                            if 목장콤보박스.get() != "선택해주세요":
                                if 직분콤보박스.get() == "목회자":
                                    테이블["목장"] = "목사님가정"
                                elif 직분콤보박스.get() == "권사":
                                    테이블["목장"] = "전도사님가정"
                                elif 직분콤보박스.get() == "청년":
                                    테이블["목장"] = "청년목장"
                                else:
                                    테이블["목장"] = 목장콤보박스.get()
                                if 년콤보박스.get() != "":
                                    테이블["생년월일"] += 년콤보박스.get() + "-"
                                    테이블["교인코드"] += 년콤보박스.get()
                                    if 월콤보박스.get() != "":
                                        테이블["생년월일"] += 월콤보박스.get() + "-"
                                        테이블["교인코드"] += 월콤보박스.get()
                                        if 일콤보박스.get() != "":
                                            테이블["생년월일"] += 일콤보박스.get()
                                            테이블["교인코드"] += 일콤보박스.get()
                                            if 테이블["가족"] != "":
                                                테이블["가족"] = 테이블["가족"][0:len(테이블["가족"]) - 2]
                                            else:
                                                테이블["가족"] = "없음"
                                            if 주소입력.get(1.0, tk.END).rstrip() != "":
                                                테이블["주소"] = 주소입력.get(1.0, tk.END).rstrip()
                                            else:
                                                테이블["주소"] = "없음"
                                            if 목장지기여부.get() == "":
                                                테이블["목장지기여부"] = "N"
                                            else:
                                                테이블["목장지기여부"] = 목장지기여부.get()
                                            if 테이블["침례세례일"] == "":
                                                테이블["침례세례일"] = "없음"
                                            if 테이블["침례세례받은교회"] == "":
                                                테이블["침례세례받은교회"] = "없음"
                                            if 테이블["침례세례목사"] == "":
                                                테이블["침례세례목사"] = "없음"

                                            sql = "SELECT * FROM " + DB테이블 + " WHERE 휴대전화번호 = '" + 테이블[
                                                "휴대전화번호"] + "';"
                                            code = cursor.execute(sql)
                                            if code == 0:
                                                sql = "INSERT INTO `" + DB테이블 + "`(교인코드, 이름, 휴대전화번호, 목장, 직분, 생년월일, 가족, 주소, 목장지기여부, 집전화번호, 침례세례일, 침례세례받은교회, 침례세례목사, 개인정보활용동의) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);"
                                                cursor.execute(sql, tuple(list(테이블.values())))
                                                church_db.commit()
                                                main.deiconify()
                                                info_input_Manual_.destroy()
                                            else:
                                                mb.showerror("오류", "존재하는 정보입니다.")
                                        else:
                                            mb.showerror("오류", "생년월일을 선택해주세요.")
                                    else:
                                        mb.showerror("오류", "생년월일을 선택해주세요.")
                                else:
                                    mb.showerror("오류", "생년월일을 선택해주세요.")
                            else:
                                mb.showerror("오류", "직분을 선택해주세요.")
                        else:
                            mb.showerror("오류", "목장을 선택해주세요.")
                    else:
                        mb.showerror("오류", "이름을 입력해주세요.")
                else:
                    mb.showerror("오류", "개인정보활용동의를 해주세요.")

            def input_stop():
                MsgBox = mb.askyesno('종료하시겠습니까?', '작성중에 종료한다면 저장되지 않습니다.', icon='warning')
                if MsgBox is True:
                    main.deiconify()
                    info_input_Manual_.destroy()

            def 동의사항():
                동의확인 = pg.confirm(
                    "개인정보보호법에 근거하여 인천중앙침례교회는 아래와 같이 성도님의 개인정보 수집·활용 동의를 받고자 합니다.\n\n1. 개인정보의 수집·활용 목적 : 인천중앙침례교회 교인 관리 및 전반적인 행정\n\n2. 이용하는 개인정보의 항목\n1) 이름\n2) 휴대폰번호, 집전화번호\n3) 집주소\n4) 생년월일\n5) 침례(세례)일, 침례(세례)받은 교회, 주례목사명\n6) 가족명\n\n3. 개인정보의 보유 및 활용기간 : 등록일로부터 제적일까지\n\n4. 성도님들은 개인정보 수집·활용에 동의하지 않으실 수 있습니다.\n\n5. 개인정보 수집·활용 동의하신 분에 한해서 교적부를 활용한 행정에 대해 지원을 받을 수 있습니다.",
                    title="개인정보활용동의", buttons=["동의", "비동의"])
                if 동의확인 == "동의":
                    mb.showinfo("동의", "개인정보활용에 동의하셨습니다.")
                    동의버튼['state'] = tk.DISABLED
                    동의버튼['text'] = "동의하셨습니다."
                    테이블["개인정보활용동의"] = "Y"
                else:
                    mb.showwarning("경고", "개인정보활용동의를 허락하지 않을 시 본 프로그램 사용이 불가능합니다.")

            def 가족추가():
                가족추가입력 = pg.prompt(title="가족 구성원 추가", text="한 명의 이름만 입력해주세요.")
                if type(가족추가입력) == str and 가족추가입력 != "":
                    테이블["가족"] += 가족추가입력
                    mb.showinfo("알람", "입력된 구성원:" + 테이블["가족"])
                    테이블["기족"] += ", "

            def 가족삭제():
                테이블["가족"] = ''
                mb.showinfo("알람", "가족 구성원 입력이 초기화 되었습니다.")

            def 침례세례():
                def 정지():
                    MsgBox = mb.askyesno('종료하시겠습니까?', '작성중에 종료한다면 저장되지 않습니다.', icon='warning')
                    if MsgBox is True:
                        침례세례창.destroy()

                def 침례세례저장():
                    MsgBox = mb.askyesno('저장하시겠습니까?', '모두 올바르게 입력하시고 저장해주세요.')
                    if MsgBox is True:
                        년월일 = ""
                        if 침례세례년.get() != '':
                            년월일 += 침례세례년.get()
                            if 침례세례달.get() != '':
                                년월일 += "-"
                                년월일 += 침례세례달.get()
                                if 침례세례일.get() != '':
                                    년월일 += "-"
                                    년월일 += 침례세례일.get()
                        else:
                            년월일 = "없음"
                        테이블["침례세례일"] = 년월일
                        if 침례세례교회입력.get() != '':
                            테이블["침례세례받은교회"] = 침례세례교회입력.get()
                        if 침례세례목사입력.get() != '':
                            테이블["침례세례목사"] = 침례세례교회입력.get()

                침례세례창 = tk.Toplevel(info_input_Manual_)
                침례세례창.title("침례(세례)")
                침례세례창.geometry("300x130")
                침례세례창.resizable(False, False)
                침례세례창.iconbitmap('assets/church_icon.ico')

                침례세례틀 = tk.Frame(침례세례창, padx=2, pady=2)
                침례세례라벨 = tk.Label(침례세례틀, text="침례(세례)일:", font=라벨폰트)
                침례세례라벨.grid(row=0, column=0)

                침례세례년 = Combobox(침례세례틀, width=4, state="readonly")
                침례세례년.grid(row=0, column=1)
                침례세례년['values'] = 년선택
                침례세례년.current(0)
                침례세례년도 = tk.Label(침례세례틀, text="년", font=라벨폰트)
                침례세례년도.grid(row=0, column=2)

                침례세례달 = Combobox(침례세례틀, width=2, state="readonly")
                침례세례달.grid(row=0, column=3)
                침례세례달['values'] = 월선택
                침례세례달.current(0)
                침례세례월 = tk.Label(침례세례틀, text="월", font=라벨폰트)
                침례세례월.grid(row=0, column=4)

                침례세례일 = Combobox(침례세례틀, width=2, state="readonly")
                침례세례일.grid(row=0, column=5)
                침례세례일['values'] = 일선택
                침례세례일.current(0)
                침례세례날 = tk.Label(침례세례틀, text="일", font=라벨폰트)
                침례세례날.grid(row=0, column=6)
                침례세례틀.grid(row=0, column=0, sticky="w")

                침례세례교회틀 = tk.Frame(침례세례창, padx=2, pady=2)
                침례세례교회 = tk.Label(침례세례교회틀, text="침례(세례)교회:", font=라벨폰트)
                침례세례교회.grid(row=0, column=0)
                침례세례교회입력 = tk.Entry(침례세례교회틀)
                침례세례교회입력.grid(row=0, column=1)
                침례세례교회틀.grid(row=1, column=0, sticky="w")

                침례세례목사틀 = tk.Frame(침례세례창, padx=2, pady=2)
                침례세례목사 = tk.Label(침례세례목사틀, text="침례(세례)목사:", font=라벨폰트)
                침례세례목사.grid(row=0, column=0)
                침례세례목사입력 = tk.Entry(침례세례목사틀)
                침례세례목사입력.grid(row=0, column=1)
                침례세례목사틀.grid(row=2, column=0, sticky="w")

                침례세례저장버튼 = tk.Button(침례세례창, text="저장", font=버튼폰트, command=침례세례저장)
                침례세례저장버튼.grid(row=3, column=0, padx=2, pady=2)

                침례세례창.protocol("WM_DELETE_WINDOW", 정지)

            main.withdraw()
            info_input_Manual_ = tk.Toplevel(main)
            info_input_Manual_.title("수동으로 정보 입력")
            info_input_Manual_.geometry("280x490")
            info_input_Manual_.resizable(False, False)
            info_input_Manual_.iconbitmap('assets/church_icon.ico')

            필수 = tk.Label(info_input_Manual_, text="* 는 필수 입력 사항입니다.", font=주석폰트, fg='red', padx=2, pady=2)
            필수.grid(row=0, column=0, sticky="w")

            동의틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            동의 = tk.Label(동의틀, text="*개인정보활용동의", font=라벨폰트)
            동의.grid(row=0, column=0)
            동의버튼 = tk.Button(동의틀, text="확인", font=버튼폰트, command=동의사항)
            동의버튼.grid(row=0, column=1)
            동의틀.grid(row=1, column=0, sticky="w")

            이름틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            이름 = tk.Label(이름틀, text="*이름:", font=라벨폰트)
            이름.grid(row=0, column=0)
            이름입력 = tk.Entry(이름틀)
            이름입력.grid(row=0, column=1)
            이름틀.grid(row=2, column=0, sticky="w")

            휴대전화번호틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            휴대전화번호 = tk.Label(휴대전화번호틀, text="휴대전화번호:", font=라벨폰트)
            휴대전화번호.grid(row=0, column=0)
            휴대전화번호입력 = tk.Entry(휴대전화번호틀)
            휴대전화번호입력.grid(row=0, column=1)
            휴대전화번호틀.grid(row=3, column=0, sticky="w")

            집전화번호틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            집전화번호 = tk.Label(집전화번호틀, text="집전화번호:", font=라벨폰트)
            집전화번호.grid(row=0, column=0)
            집전화번호입력 = tk.Entry(집전화번호틀)
            집전화번호입력.grid(row=0, column=1)
            집전화번호틀.grid(row=4, column=0, sticky="w")

            목장틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            목장 = tk.Label(목장틀, text="*목장:", font=라벨폰트)
            목장.grid(row=0, column=0)
            목장콤보박스 = Combobox(목장틀, state="readonly")
            목장콤보박스['values'] = 목장선택
            목장콤보박스.current(0)
            목장콤보박스.grid(row=0, column=1)
            목장틀.grid(row=5, column=0, sticky="w")

            직분틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            직분 = tk.Label(직분틀, text="*직분:", font=라벨폰트)
            직분.grid(row=0, column=0)
            직분콤보박스 = Combobox(직분틀, state="readonly")
            직분콤보박스['values'] = 직분선택
            직분콤보박스.current(0)
            직분콤보박스.grid(row=0, column=1)
            직분틀.grid(row=6, column=0, sticky="w")

            생일틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            생년월일 = tk.Label(생일틀, text="*생일:", font=라벨폰트)
            생년월일.grid(row=0, column=0)

            년콤보박스 = Combobox(생일틀, width=4, state="readonly")
            년콤보박스.grid(row=0, column=1)
            년콤보박스['values'] = 년선택
            년 = tk.Label(생일틀, text="년", font=라벨폰트)
            년.grid(row=0, column=2)

            월콤보박스 = Combobox(생일틀, width=2, state="readonly")
            월콤보박스.grid(row=0, column=3)
            월콤보박스['values'] = 월선택
            월 = tk.Label(생일틀, text="월", font=라벨폰트)
            월.grid(row=0, column=4)

            일콤보박스 = Combobox(생일틀, width=2, state="readonly")
            일콤보박스.grid(row=0, column=5)
            일콤보박스['values'] = 일선택
            일 = tk.Label(생일틀, text="일", font=라벨폰트)
            일.grid(row=0, column=6)
            생일틀.grid(row=7, column=0, sticky="w")

            가족틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            가족 = tk.Label(가족틀, text="가족:", font=라벨폰트)
            가족.grid(row=0, column=0)
            가족입력버튼 = tk.Button(가족틀, text="추가", font=버튼폰트, command=가족추가)
            가족입력버튼.grid(row=0, column=1)
            가족입력버튼 = tk.Button(가족틀, text="삭제", font=버튼폰트, command=가족삭제)
            가족입력버튼.grid(row=0, column=2)
            가족틀.grid(row=8, column=0, sticky="w")

            주소틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            주소예시 = tk.Label(주소틀, text="현 거주지를 입력해주세요.\nex) XX시 XX구 XX동 XX아파트 X동 XXX호", font=주석폰트, fg='Gray')
            주소예시.grid(row=0, column=0, columnspan=5)
            주소 = tk.Label(주소틀, text="*주소:", font=라벨폰트)
            주소.grid(row=1, column=0, sticky="n")
            주소입력 = tk.Text(주소틀, height=1, width=25)
            주소입력.grid(row=1, column=1, ipadx=20, ipady=30)
            주소틀.grid(row=9, column=0, sticky="w")

            목장지기틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            목장지기여부 = tk.StringVar()
            목장지기 = tk.Label(목장지기틀, text="*목장지기 여부:", font=라벨폰트)
            목장지기.grid(row=0, column=0)
            목장지기맞다 = tk.Radiobutton(목장지기틀, text="맞다", value="Y", variable=목장지기여부)
            목장지기맞다.grid(row=0, column=1)
            목장지기아니다 = tk.Radiobutton(목장지기틀, text="아니다", value="N", variable=목장지기여부)
            목장지기아니다.grid(row=0, column=2)
            목장지기틀.grid(row=10, column=0, sticky="w")

            침례세례버튼 = tk.Button(info_input_Manual_, text="침례(세례)를 했을 시 눌러주세요.", font=버튼폰트, command=침례세례)
            침례세례버튼.grid(row=11, column=0, padx=2, pady=2)

            저장버튼 = tk.Button(info_input_Manual_, text="저장", font=버튼폰트, command=저장)
            저장버튼.grid(row=12, column=0, padx=2, pady=2)

            info_input_Manual_.protocol("WM_DELETE_WINDOW", input_stop)

        def 수정입력():
            주석폰트 = tf.Font(family="맑은 고딕", size=8)
            라벨폰트 = tf.Font(family="맑은 고딕", size=12)
            버튼폰트 = tf.Font(family="맑은 고딕", size=10)
            테이블 = {"교인코드": "", "이름": "", "휴대전화번호": "", "목장": "",
                   "직분": "", "생년월일": "", "가족": "", "주소": "",
                   "목장지기여부": "", "집전화번호": "", "침례세례일": "",
                   "침례세례받은교회": "", "침례세례목사": "", "개인정보활용동의": ""}

            def 저장():
                if 동의버튼['state'] == tk.DISABLED:
                    테이블["개인정보활용동의"] = "Y"
                    if 이름입력.get() != '':
                        테이블["이름"] = 이름입력.get().replace(" ", "")
                        if 휴대전화번호입력.get() != '' and len(휴대전화번호입력.get()) >= 11:
                            테이블["휴대전화번호"] = 휴대전화번호입력.get()[:11]
                            filter(str.isdigit, 테이블["휴대전화번호"])
                        else:
                            테이블["휴대전화번호"] = "없음"
                        if 집전화번호입력.get() != '':
                            테이블["집전화번호"] = 집전화번호입력.get()
                        else:
                            테이블["집전화번호"] = "없음"
                        if 직분콤보박스.get() != "선택해주세요":
                            테이블["직분"] = 직분콤보박스.get()
                            if 목장콤보박스.get() != "선택해주세요":
                                if 직분콤보박스.get() == "목회자":
                                    테이블["목장"] = "목사님가정"
                                elif 직분콤보박스.get() == "권사":
                                    테이블["목장"] = "전도사님가정"
                                elif 직분콤보박스.get() == "청년":
                                    테이블["목장"] = "청년목장"
                                else:
                                    테이블["목장"] = 목장콤보박스.get()
                                if 년콤보박스.get() != "":
                                    테이블["생년월일"] += 년콤보박스.get() + "-"
                                    테이블["교인코드"] += 년콤보박스.get()
                                    if 월콤보박스.get() != "":
                                        테이블["생년월일"] += 월콤보박스.get() + "-"
                                        테이블["교인코드"] += 월콤보박스.get()
                                        if 일콤보박스.get() != "":
                                            테이블["생년월일"] += 일콤보박스.get()
                                            테이블["교인코드"] += 일콤보박스.get()
                                            if 테이블["가족"] != "":
                                                테이블["가족"] = 테이블["가족"][0:len(테이블["가족"]) - 2]
                                            else:
                                                테이블["가족"] = "없음"
                                            if 주소입력.get(1.0, tk.END).rstrip() != "":
                                                테이블["주소"] = 주소입력.get(1.0, tk.END).rstrip()
                                            else:
                                                테이블["주소"] = "없음"
                                            if 목장지기여부.get() == "":
                                                테이블["목장지기여부"] = "N"
                                            else:
                                                테이블["목장지기여부"] = 목장지기여부.get()
                                            if 테이블["침례세례일"] == "":
                                                테이블["침례세례일"] = "없음"
                                            if 테이블["침례세례받은교회"] == "":
                                                테이블["침례세례받은교회"] = "없음"
                                            if 테이블["침례세례목사"] == "":
                                                테이블["침례세례목사"] = "없음"

                                            sql = "SELECT * FROM " + DB테이블 + " WHERE 교인코드 = '" + 테이블["교인코드"] + "';"
                                            code = cursor.execute(sql)
                                            if code != 0:
                                                for 키, 밸류 in zip(list(테이블.keys()), list(테이블.values())):
                                                    테이블[키] = 밸류
                                                    sql = "UPDATE " + DB테이블 + " SET " + 키 + " = '" + 밸류 + "' WHERE 교인코드 = '" + \
                                                          테이블["교인코드"] + "';"
                                                    cursor.execute(sql)
                                                    church_db.commit()
                                                main.deiconify()
                                                info_input_Manual_.destroy()
                                            else:
                                                print("안돼")
                                        else:
                                            mb.showerror("오류", "생년월일을 선택해주세요.")
                                    else:
                                        mb.showerror("오류", "생년월일을 선택해주세요.")
                                else:
                                    mb.showerror("오류", "생년월일을 선택해주세요.")
                            else:
                                mb.showerror("오류", "직분을 선택해주세요.")
                        else:
                            mb.showerror("오류", "목장을 선택해주세요.")
                    else:
                        mb.showerror("오류", "이름을 입력해주세요.")
                else:
                    mb.showerror("오류", "개인정보활용동의를 해주세요.")

            def input_stop():
                MsgBox = mb.askyesno('종료하시겠습니까?', '작성중에 종료한다면 저장되지 않습니다.', icon='warning')
                if MsgBox is True:
                    main.deiconify()
                    info_input_Manual_.destroy()

            def 가족추가():
                가족추가입력 = pg.prompt(title="가족 구성원 추가", text="한 명의 이름만 입력해주세요.")
                if type(가족추가입력) == str and 가족추가입력 != "":
                    테이블["가족"] += 가족추가입력
                    mb.showinfo("알람", "입력된 구성원:" + 테이블["가족"])
                    테이블["기족"] += ", "

            def 가족삭제():
                테이블["가족"] = ''
                mb.showinfo("알람", "가족 구성원 입력이 초기화 되었습니다.")

            def 침례세례():
                def 정지():
                    MsgBox = mb.askyesno('종료하시겠습니까?', '작성중에 종료한다면 저장되지 않습니다.', icon='warning')
                    if MsgBox is True:
                        침례세례창.destroy()

                def 침례세례저장():
                    MsgBox = mb.askyesno('저장하시겠습니까?', '모두 올바르게 입력하시고 저장해주세요.')
                    if MsgBox is True:
                        년월일 = ""
                        if 침례세례년.get() != '':
                            년월일 += 침례세례년.get()
                            if 침례세례달.get() != '':
                                년월일 += "-"
                                년월일 += 침례세례달.get()
                                if 침례세례일.get() != '':
                                    년월일 += "-"
                                    년월일 += 침례세례일.get()
                        else:
                            년월일 = "없음"
                        테이블["침례세례일"] = 년월일
                        if 침례세례교회입력.get() != '':
                            테이블["침례세례받은교회"] = 침례세례교회입력.get()
                        if 침례세례목사입력.get() != '':
                            테이블["침례세례목사"] = 침례세례교회입력.get()

                침례세례창 = tk.Toplevel(info_input_Manual_)
                침례세례창.title("침례(세례)")
                침례세례창.geometry("300x130")
                침례세례창.resizable(False, False)
                침례세례창.iconbitmap('assets/church_icon.ico')

                침례세례틀 = tk.Frame(침례세례창, padx=2, pady=2)
                침례세례라벨 = tk.Label(침례세례틀, text="침례(세례)일:", font=라벨폰트)
                침례세례라벨.grid(row=0, column=0)

                침례세례년 = Combobox(침례세례틀, width=4, state="readonly")
                침례세례년.grid(row=0, column=1)
                침례세례년['values'] = 년선택
                침례세례년.current(0)
                침례세례년도 = tk.Label(침례세례틀, text="년", font=라벨폰트)
                침례세례년도.grid(row=0, column=2)

                침례세례달 = Combobox(침례세례틀, width=2, state="readonly")
                침례세례달.grid(row=0, column=3)
                침례세례달['values'] = 월선택
                침례세례달.current(0)
                침례세례월 = tk.Label(침례세례틀, text="월", font=라벨폰트)
                침례세례월.grid(row=0, column=4)

                침례세례일 = Combobox(침례세례틀, width=2, state="readonly")
                침례세례일.grid(row=0, column=5)
                침례세례일['values'] = 일선택
                침례세례일.current(0)
                침례세례날 = tk.Label(침례세례틀, text="일", font=라벨폰트)
                침례세례날.grid(row=0, column=6)
                침례세례틀.grid(row=0, column=0, sticky="w")

                침례세례교회틀 = tk.Frame(침례세례창, padx=2, pady=2)
                침례세례교회 = tk.Label(침례세례교회틀, text="침례(세례)교회:", font=라벨폰트)
                침례세례교회.grid(row=0, column=0)
                침례세례교회입력 = tk.Entry(침례세례교회틀)
                침례세례교회입력.grid(row=0, column=1)
                침례세례교회틀.grid(row=1, column=0, sticky="w")

                침례세례목사틀 = tk.Frame(침례세례창, padx=2, pady=2)
                침례세례목사 = tk.Label(침례세례목사틀, text="침례(세례)목사:", font=라벨폰트)
                침례세례목사.grid(row=0, column=0)
                침례세례목사입력 = tk.Entry(침례세례목사틀)
                침례세례목사입력.grid(row=0, column=1)
                침례세례목사틀.grid(row=2, column=0, sticky="w")

                침례세례저장버튼 = tk.Button(침례세례창, text="저장", font=버튼폰트, command=침례세례저장)
                침례세례저장버튼.grid(row=3, column=0, padx=2, pady=2)

                침례세례창.protocol("WM_DELETE_WINDOW", 정지)

            def 검색():
                info_input_Manual_.withdraw()
                검색입력 = pg.prompt('수정할 이름을 입력해주세요.')

                if 검색입력 is None:
                    main.deiconify()
                else:
                    sql = "SELECT * FROM " + DB테이블 + " WHERE 이름 = '" + 검색입력 + "';"
                    code = cursor.execute(sql)
                    if code == 0:
                        mb.showerror("오류", "존재하지 않는 이름입니다.")
                        검색()
                    elif code == 1:
                        for 목록 in cursor.fetchall():
                            for 키, 밸류 in zip(list(테이블.keys()), list(목록.values())):
                                테이블[키] = 밸류
                        테이블["교인코드"] = 테이블["교인코드"][:14]
                        info_input_Manual_.deiconify()
                        이름입력.insert(0, 테이블["이름"])
                        휴대전화번호입력.insert(0, 테이블["휴대전화번호"])
                        집전화번호입력.insert(0, 테이블["집전화번호"])
                        목장콤보박스.current(목장선택.index(테이블["목장"]))
                        직분콤보박스.current(직분선택.index(테이블["직분"]))
                        년콤보박스.current(년선택.index(테이블["생년월일"][0:4]))
                        월콤보박스.current(월선택.index(테이블["생년월일"][5:7]))
                        일콤보박스.current(일선택.index(테이블["생년월일"][8:10]))
                        주소입력.insert("1.0", 테이블["주소"])
                        if 테이블["목장지기여부"] == 'Y':
                            목장지기맞다.select()
                        else:
                            목장지기아니다.select()

            main.withdraw()
            info_input_Manual_ = tk.Toplevel(main)
            info_input_Manual_.title("정보 수정")
            info_input_Manual_.geometry("280x490")
            info_input_Manual_.resizable(False, False)
            info_input_Manual_.iconbitmap('assets/church_icon.ico')

            필수 = tk.Label(info_input_Manual_, text="* 는 필수 입력 사항입니다.", font=주석폰트, fg='red', padx=2, pady=2)
            필수.grid(row=0, column=0, sticky="w")

            동의틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            동의 = tk.Label(동의틀, text="*개인정보활용동의", font=라벨폰트)
            동의.grid(row=0, column=0)
            동의버튼 = tk.Button(동의틀, text="동의하셨습니다", font=버튼폰트)
            동의버튼['state'] = tk.DISABLED
            동의버튼.grid(row=0, column=1)
            동의틀.grid(row=1, column=0, sticky="w")

            이름틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            이름 = tk.Label(이름틀, text="*이름:", font=라벨폰트)
            이름.grid(row=0, column=0)
            이름입력 = tk.Entry(이름틀)
            이름입력.grid(row=0, column=1)
            이름틀.grid(row=2, column=0, sticky="w")

            휴대전화번호틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            휴대전화번호 = tk.Label(휴대전화번호틀, text="휴대전화번호:", font=라벨폰트)
            휴대전화번호.grid(row=0, column=0)
            휴대전화번호입력 = tk.Entry(휴대전화번호틀)
            휴대전화번호입력.grid(row=0, column=1)
            휴대전화번호틀.grid(row=3, column=0, sticky="w")

            집전화번호틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            집전화번호 = tk.Label(집전화번호틀, text="집전화번호:", font=라벨폰트)
            집전화번호.grid(row=0, column=0)
            집전화번호입력 = tk.Entry(집전화번호틀)
            집전화번호입력.grid(row=0, column=1)
            집전화번호틀.grid(row=4, column=0, sticky="w")

            목장틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            목장 = tk.Label(목장틀, text="*목장:", font=라벨폰트)
            목장.grid(row=0, column=0)
            목장콤보박스 = Combobox(목장틀, state="readonly")
            목장콤보박스['values'] = 목장선택
            목장콤보박스.grid(row=0, column=1)
            목장틀.grid(row=5, column=0, sticky="w")

            직분틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            직분 = tk.Label(직분틀, text="*직분:", font=라벨폰트)
            직분.grid(row=0, column=0)
            직분콤보박스 = Combobox(직분틀, state="readonly")
            직분콤보박스['values'] = 직분선택
            직분콤보박스.grid(row=0, column=1)
            직분틀.grid(row=6, column=0, sticky="w")

            생일틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            생년월일 = tk.Label(생일틀, text="*생일:", font=라벨폰트)
            생년월일.grid(row=0, column=0)

            년콤보박스 = Combobox(생일틀, width=4, state="readonly")
            년콤보박스.grid(row=0, column=1)
            년콤보박스['values'] = 년선택
            년 = tk.Label(생일틀, text="년", font=라벨폰트)
            년.grid(row=0, column=2)

            월콤보박스 = Combobox(생일틀, width=2, state="readonly")
            월콤보박스.grid(row=0, column=3)
            월콤보박스['values'] = 월선택
            월 = tk.Label(생일틀, text="월", font=라벨폰트)
            월.grid(row=0, column=4)

            일콤보박스 = Combobox(생일틀, width=2, state="readonly")
            일콤보박스.grid(row=0, column=5)
            일콤보박스['values'] = 일선택
            일 = tk.Label(생일틀, text="일", font=라벨폰트)
            일.grid(row=0, column=6)
            생일틀.grid(row=7, column=0, sticky="w")

            가족틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            가족 = tk.Label(가족틀, text="가족:", font=라벨폰트)
            가족.grid(row=0, column=0)
            가족입력버튼 = tk.Button(가족틀, text="추가", font=버튼폰트, command=가족추가)
            가족입력버튼.grid(row=0, column=1)
            가족입력버튼 = tk.Button(가족틀, text="삭제", font=버튼폰트, command=가족삭제)
            가족입력버튼.grid(row=0, column=2)
            가족틀.grid(row=8, column=0, sticky="w")

            주소틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            주소예시 = tk.Label(주소틀, text="현 거주지를 입력해주세요.\nex) XX시 XX구 XX동 XX아파트 X동 XXX호", font=주석폰트, fg='Gray')
            주소예시.grid(row=0, column=0, columnspan=5)
            주소 = tk.Label(주소틀, text="*주소:", font=라벨폰트)
            주소.grid(row=1, column=0, sticky="n")
            주소입력 = tk.Text(주소틀, height=1, width=25)
            주소입력.grid(row=1, column=1, ipadx=20, ipady=30)
            주소틀.grid(row=9, column=0, sticky="w")

            목장지기틀 = tk.Frame(info_input_Manual_, padx=2, pady=2)
            목장지기여부 = tk.StringVar()
            목장지기 = tk.Label(목장지기틀, text="*목장지기 여부:", font=라벨폰트)
            목장지기.grid(row=0, column=0)
            목장지기맞다 = tk.Radiobutton(목장지기틀, text="맞다", value="Y", variable=목장지기여부)
            목장지기맞다.grid(row=0, column=1)
            목장지기아니다 = tk.Radiobutton(목장지기틀, text="아니다", value="N", variable=목장지기여부)
            목장지기아니다.grid(row=0, column=2)
            목장지기틀.grid(row=10, column=0, sticky="w")

            침례세례버튼 = tk.Button(info_input_Manual_, text="침례(세례)를 했을 시 눌러주세요.", font=버튼폰트, command=침례세례)
            침례세례버튼.grid(row=11, column=0, padx=2, pady=2)

            저장버튼 = tk.Button(info_input_Manual_, text="저장", font=버튼폰트, command=저장)
            저장버튼.grid(row=12, column=0, padx=2, pady=2)

            검색()

            info_input_Manual_.protocol("WM_DELETE_WINDOW", input_stop)

        def 종료():
            cursor.close()
            church_db.commit()
            church_db.close()
            main.destroy()
            sys.exit()

        # noinspection PyUnusedLocal
        def 메일보내기():
            오늘날짜 = str(dt.datetime.today().date())
            try:
                엑셀 = xl.load_workbook("data/" + 오늘날짜 + ".xlsx")
                os.remove("data/" + 오늘날짜 + ".xlsx")
                엑셀 = xl.Workbook()
            except:
                엑셀 = xl.Workbook()
            시트 = 엑셀.active

            try:
                교인이름 = []
                명령 = "SELECT 이름, 교인코드 FROM " + DB테이블 + ";"
                cursor.execute(명령)
                for 값 in cursor.fetchall():
                    교인이름.append([값["교인코드"], 값["이름"]])
                교인이름 = dict(교인이름)
                명령 = "SELECT * FROM " + 누계테이블 + " WHERE 날짜 LIKE '" + 오늘날짜 + "';"
                cursor.execute(명령)
                data = cursor.fetchall()
                시트.append(["이름", "날짜", "금액", "형식"])
                일일누계 = 0
                for val in data:
                    시트.append([교인이름[val["교인코드"]], val["날짜"], str(format(int(val["금액"]), ",")), 헌금형식[val["형식"]]])
                    일일누계 += int(val["금액"])
                시트.append(["누계 : " + str(format(일일누계) + "\\")])

                엑셀.save("data/" + 오늘날짜 + ".xlsx")

                msg = MIMEMultipart()
                msg['Subject'] = 오늘날짜 + "자 헌금누계입니다."
                msg.attach(MIMEText(오늘날짜, 'plain'))
                attachment = open("data/" + 오늘날짜 + ".xlsx", 'rb')
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', "attachment; filename= " + "data/" + 오늘날짜 + ".xlsx")
                msg.attach(part)
                s.sendmail(보내는메일, 받는메일, msg.as_string())

                msg = MIMEMultipart()
                msg['Subject'] = 오늘날짜 + "자 출입명부입니다."
                msg.attach(MIMEText(오늘날짜 + " 출입명부", 'plain'))
                attachment = open("data/" + 오늘날짜 + " 출입명부.xlsx", 'rb')
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', "attachment; filename= " + "data/" + 오늘날짜 + " 출입명부.xlsx")
                msg.attach(part)
                s.sendmail(보내는메일, 받는메일, msg.as_string())
                s.quit()

                mb.showinfo("완료", "이번 주 헌금누계, 출입명부가 보내졌습니다.")
            except:
                mb.showerror("오류", "오류가 발생했습니다.")

        def 헌금입력():
            라벨폰트 = tf.Font(family="맑은 고딕", size=15)

            def 입력중지():
                MsgBox = mb.askyesno('종료하시겠습니까?', '작성중에 종료한다면 저장되지 않습니다.', icon='warning')
                if MsgBox is True:
                    main.deiconify()
                    헌금입력창.destroy()

            # noinspection PyUnusedLocal
            def 검색(self):
                try:
                    코드 = 입력받기.get()[:len(입력받기.get()) - 2]
                    형식 = int(입력받기.get()[len(입력받기.get()) - 1:])
                    액수 = 헌금.get()
                    날짜 = dt.datetime.today().strftime("%F")
                    입력받기.delete(0, 'end')
                    명령 = "SELECT 이름 FROM " + DB테이블 + " WHERE 교인코드 = '" + 코드 + "';"
                    값 = cursor.execute(명령)
                    if 값 == 1:
                        형식콤보.current(형식 - 1)
                        이름['text'] = "이름 : "
                        이름['text'] += cursor.fetchone()["이름"]
                        값 = (코드, 액수, 형식, 날짜)
                        if 헌금.get() != "":
                            헌금.delete(0, 'end')
                            #   형식 :
                            #       1 : 십일조
                            #       2 : 주일헌금
                            #       3 : 건축헌금
                            #       4 : 선교헌금
                            #       5 : 감사헌금
                            #       6 : 특별헌금
                            명령 = "INSERT INTO `" + 누계테이블 + "` (교인코드, 금액, 형식, 날짜) VALUES (%s, %s, %s, %s);"
                            cursor.execute(명령, 값)
                            church_db.commit()
                            헌금.focus()
                        else:
                            헌금.focus()
                    elif 값 == 0:
                        mb.showerror("오류", "존재하지 않는 코드입니다.")
                        입력받기['text'] = ""
                        입력받기.focus()
                except:
                    mb.showerror("오류", "잘못된 입력값입니다.")
                    입력받기.focus()

            # noinspection PyUnusedLocal
            def 헌금받기(self):
                입력받기.focus()

            def 저장():
                if 이름['text'] != "이름 : " and 헌금.get() != "":
                    명령 = "SELECT 교인코드 FROM " + DB테이블 + " WHERE 이름 = '" + 이름['text'][5:] + "';"
                    cursor.execute(명령)
                    교인코드 = list(cursor.fetchall()[0].values())[0]
                    금액 = 헌금.get()
                    형식 = str(형식콤보.current() + 1)
                    날짜 = dt.datetime.today().strftime("%F")
                    값 = (교인코드, 금액, 형식, 날짜)
                    명령 = "INSERT INTO `" + 누계테이블 + "` (교인코드, 금액, 형식, 날짜) VALUES (%s, %s, %s, %s);"
                    cursor.execute(명령, 값)
                    church_db.commit()
                    이름['text'] = "이름 : "
                    헌금.delete(0, 'end')
                    입력받기.focus()
                    mb.showinfo("저장", "저장되었습니다.")
                else:
                    mb.showerror("오류", "저장이 되지 않았습니다.")

            main.withdraw()
            헌금입력창 = tk.Toplevel(main)
            헌금입력창.title("헌금 입력")
            헌금입력창.geometry("550x140")
            헌금입력창.resizable(False, False)
            헌금입력창.iconbitmap('assets/church_icon.ico')

            입력라벨 = tk.Label(헌금입력창, text="교인코드 :", font=라벨폰트)
            입력라벨.grid(row=0, column=0, padx=10, pady=10, sticky="w")
            입력받기 = tk.Entry(헌금입력창, font=라벨폰트, width=25)
            입력받기.grid(row=0, column=1, padx=10, pady=10, sticky="w", columnspan=2)
            입력받기.bind("<Return>", 검색)
            입력받기.focus()
            저장버튼 = tk.Button(헌금입력창, text="저장", font=라벨폰트, command=저장)
            저장버튼.grid(row=0, column=3, padx=10, pady=10, sticky="w")

            이름 = tk.Label(헌금입력창, text="이름 : ", font=라벨폰트)
            이름.grid(row=1, column=0, padx=10, pady=10, sticky="w")

            헌금라벨 = tk.Label(헌금입력창, text="헌금 :", font=라벨폰트)
            헌금라벨.grid(row=1, column=1, padx=10, pady=10, sticky="w")
            헌금 = tk.Entry(헌금입력창, font=라벨폰트, width=15)
            헌금.grid(row=1, column=2, padx=10, pady=10, sticky="w")
            헌금.bind("<Return>", 헌금받기)
            입력받기.bind("<FocusIn>", 헌금.delete(0, 'end'))

            형식콤보 = Combobox(헌금입력창, width=7, state="readonly")
            형식콤보.grid(row=1, column=3, padx=10, pady=10, sticky="w")
            형식콤보['values'] = ["십일조", "주일헌금", "건축헌금", "선교헌금", "감사헌금", "특별헌금"]
            형식콤보.current(1)

            헌금입력창.protocol("WM_DELETE_WINDOW", 입력중지)

        def 누계확인():
            폰트 = tf.Font(family="맑은 고딕", size=15)
            리스트 = []

            def 입력중지():
                MsgBox = mb.askyesno('종료하시겠습니까?', '작성중에 종료한다면 저장되지 않습니다.', icon='warning')
                if MsgBox is True:
                    main.deiconify()
                    누계확인창.destroy()

            def 정보삭제():
                for val in 누계표.get_children():
                    누계표.delete(val)

            # noinspection PyUnusedLocal
            def 개인선택(self):
                누계확인창.focus()
                if 선택콤보박스.get() != "교회":
                    명령 = "SELECT 가족 FROM " + DB테이블 + " WHERE 교인코드 = '" + 선택콤보박스.get()[-22:] + "';"
                    cursor.execute(명령)
                    가족콤보박스['values'] = ["선택"] + list(cursor.fetchall()[0].values())[0].replace(' ', '').split(',')
                    가족콤보박스.current(0)
                    년콤보박스.current(0)
                    월콤보박스.current(0)
                    주일콤보박스.current(0)
                    헌금형식콤보박스.current(0)
                elif 선택콤보박스.get() == "교회":
                    가족콤보박스['values'] = ["없음"]
                    가족콤보박스.current(0)
                    년콤보박스.current(0)
                    월콤보박스.current(0)
                    주일콤보박스.current(0)
                    헌금형식콤보박스.current(0)
                else:
                    가족콤보박스['values'] = ["선택"]
                    가족콤보박스.current(0)
                    년콤보박스.current(0)
                    월콤보박스.current(0)
                    주일콤보박스.current(0)
                    헌금형식콤보박스.current(0)
                누계선택("temp")

            # noinspection PyUnusedLocal
            def 가족선택(self):
                누계확인창.focus()
                if 가족콤보박스.get() != "선택해주세요" and 가족콤보박스.get() != "" and 선택콤보박스.get() != "교회":
                    명령 = "SELECT 가족 FROM " + DB테이블 + " WHERE 이름 = '" + 가족콤보박스.get() + "';"
                    값 = cursor.execute(명령)
                    if 값 != 0:
                        가족콤보박스['values'] = ["선택"] + list(cursor.fetchall()[0].values())[0].replace(' ', '').split(',')
                        for 인덱스, 개인 in enumerate(리스트):
                            if 개인[:len(가족콤보박스.get())] == 가족콤보박스.get():
                                선택콤보박스.current(인덱스)
                    else:
                        mb.showerror("오류", "가족의 정보가 입력되어있지 않습니다.")
                    가족콤보박스.current(0)
                    년콤보박스.current(0)
                    월콤보박스.current(0)
                    주일콤보박스.current(0)
                    헌금형식콤보박스.current(0)
                누계선택("temp")

            def 교회확인():
                선택콤보박스.current(0)
                년콤보박스.current(0)
                월콤보박스.current(0)
                주일콤보박스.current(0)
                헌금형식콤보박스.current(0)
                개인선택("temp")

            # noinspection PyUnusedLocal
            def 누계선택(self):
                정보삭제()
                누계확인창.focus()
                IV = []
                명령 = "SELECT * FROM " + 누계테이블 + ";"
                cursor.execute(명령)
                data = cursor.fetchall()
                총누계값, 누계값 = 0, 0
                for val in data:
                    if 선택콤보박스.get()[-교인코드길이:] == val["교인코드"]:
                        IV.append([val["날짜"], val["금액"], 헌금형식[val["형식"]]])
                        총누계값 += int(val["금액"])
                    elif 선택콤보박스.get() == "교회":
                        IV.append([val["날짜"], val["금액"], 헌금형식[val["형식"]]])
                        if val["날짜"][:4] == 년콤보박스.get():
                            누계값 += int(val["금액"])
                        총누계값 += int(val["금액"])

                for val in IV:
                    년모두, 월모두, 주일모두, 형식모두, 년일치, 월일치, 주일일치, 형식일치 = \
                        년콤보박스.get() == "모두", 월콤보박스.get() == "모두", 주일콤보박스.get() == "모두", 헌금형식콤보박스.get() == "모두", \
                        val[0][:4] == 년콤보박스.get(), val[0][5:7] == 월콤보박스.get(), val[0][8:] == 주일콤보박스.get(), val[
                            2] == 헌금형식콤보박스.get()
                    개인조건 = 년모두 and 월모두 and 주일모두 and 형식모두 or \
                           년모두 and 월모두 and 주일모두 and 형식일치 or 년모두 and 월모두 and 주일일치 and 형식모두 or \
                           년모두 and 월일치 and 주일모두 and 형식모두 or 년일치 and 월모두 and 주일모두 and 형식모두 or \
                           년모두 and 월모두 and 주일일치 and 형식일치 or 년모두 and 월일치 and 주일일치 and 형식모두 or \
                           년일치 and 월일치 and 주일모두 and 형식모두 or 년일치 and 월모두 and 주일모두 and 형식일치 or \
                           년모두 and 월일치 and 주일모두 and 형식일치 or 년일치 and 월모두 and 주일일치 and 형식모두 or \
                           년일치 and 월일치 and 주일일치 and 형식모두 or 년모두 and 월일치 and 주일일치 and 형식일치 or \
                           년일치 and 월모두 and 주일일치 and 형식일치 or 년일치 and 월일치 and 주일모두 and 형식일치 or \
                           년일치 and 월일치 and 주일일치 and 형식모두 or 년일치 and 월일치 and 주일일치 and 형식일치
                    if 개인조건:
                        누계표.insert('', 'end', values=(val[0], str(format(int(val[1]), ",")), val[2]))
                        누계값 += int(val[1])
                누계라벨['text'] = "누계 : " + str(format(누계값, ",")) + "\\ | 총누계 : " + str(format(총누계값, ",")) + "\\"

            main.withdraw()
            누계확인창 = tk.Toplevel(main)
            누계확인창.title("누계 확인")
            누계확인창.geometry("1150x610")
            누계확인창.resizable(False, False)
            누계확인창.iconbitmap('assets/church_icon.ico')

            교회버튼 = tk.Button(누계확인창, command=교회확인, text="교회", font=폰트)
            교회버튼.grid(row=0, column=0, padx=10, pady=10, sticky="w")

            명령 = "SELECT 이름, 교인코드 FROM " + DB테이블 + ";"
            cursor.execute(명령)
            for 값 in cursor.fetchall():
                리스트값 = ""
                for i in list(값.values()):
                    리스트값 += i
                리스트.append(리스트값)
            for 인덱스, 값 in enumerate(리스트):
                리스트[인덱스] = 값[:-교인코드길이] + " : " + 값[-교인코드길이:]
            리스트.insert(0, "교회")

            선택콤보박스 = Combobox(누계확인창, state="readonly", font=폰트, width=30)
            선택콤보박스['values'] = 리스트
            선택콤보박스.grid(row=0, column=1, pady=10, sticky="w")
            선택콤보박스.current(0)
            선택콤보박스.bind("<<ComboboxSelected>>", 개인선택)

            년선택라벨 = tk.Label(누계확인창, text="년 :", font=폰트)
            년선택라벨.grid(row=0, column=2, pady=10, sticky="w")

            년콤보박스 = Combobox(누계확인창, state="readonly", font=폰트, width=6)
            년콤보박스.grid(row=0, column=3, pady=10, sticky="w")
            년콤보박스['values'] = ("모두",) + 년선택
            년콤보박스.current(0)
            년콤보박스.bind("<<ComboboxSelected>>", 누계선택)

            월라벨 = tk.Label(누계확인창, text="월 :", font=폰트)
            월라벨.grid(row=0, column=4, pady=10, sticky="w")

            월콤보박스 = Combobox(누계확인창, state="readonly", font=폰트, width=6)
            월콤보박스.grid(row=0, column=5, pady=10, sticky="w")
            월콤보박스['values'] = ("모두",) + 월선택
            월콤보박스.current(0)
            월콤보박스.bind("<<ComboboxSelected>>", 누계선택)

            주일형식라벨 = tk.Label(누계확인창, text="주일 :", font=폰트)
            주일형식라벨.grid(row=0, column=6, pady=10, sticky="w")

            주일콤보박스 = Combobox(누계확인창, state="readonly", font=폰트, width=6)
            주일콤보박스.grid(row=0, column=7, pady=10, sticky="w")
            주일콤보박스['values'] = ("모두",) + 일선택
            주일콤보박스.current(0)
            주일콤보박스.bind("<<ComboboxSelected>>", 누계선택)

            헌금형식라벨 = tk.Label(누계확인창, text="형식 :", font=폰트)
            헌금형식라벨.grid(row=0, column=8, pady=10, sticky="w")

            헌금형식콤보박스 = Combobox(누계확인창, state="readonly", font=폰트, width=8)
            헌금형식콤보박스.grid(row=0, column=9, pady=10, sticky="w")
            헌금형식콤보박스['values'] = ["모두", "십일조", "주일헌금", "건축헌금", "선교헌금", "감사헌금", "특별헌금"]
            헌금형식콤보박스.current(0)
            헌금형식콤보박스.bind("<<ComboboxSelected>>", 누계선택)

            누계틀 = tk.Frame(누계확인창)
            누계표 = ttk.Treeview(누계틀, columns=["Date", "Amount", "Type"], height=23, show="headings")
            누계표.pack(side="left")

            누계표.heading("Date", text="날짜")
            누계표.heading("Amount", text="금액")
            누계표.heading("Type", text="형식")
            누계표.column("Date", width=370, anchor="e")
            누계표.column("Amount", width=370, anchor="e")
            누계표.column("Type", width=370, anchor="e")

            스크롤바 = ttk.Scrollbar(누계틀)
            스크롤바.pack(side="right", fill="y")
            스크롤바["command"] = 누계표.yview

            누계틀.grid(row=1, column=0, padx=10, sticky="w", columnspan=10)

            누계라벨 = tk.Label(누계확인창, text="누계 : 0\\ | 총누계 : 0\\", font=폰트)
            누계라벨.grid(row=2, column=0, padx=10, pady=10, sticky="w", columnspan=15)

            가족라벨 = tk.Label(누계확인창, text="가족 :", font=폰트)
            가족라벨.grid(row=2, column=8, pady=10, sticky="w")

            가족콤보박스 = Combobox(누계확인창, state="readonly", font=폰트, width=6)
            가족콤보박스.grid(row=2, column=9, pady=10, sticky="w")
            가족콤보박스['values'] = ["선택"]
            가족콤보박스.current(0)
            가족콤보박스.bind("<<ComboboxSelected>>", 가족선택)

            개인선택("temp")

            누계확인창.protocol("WM_DELETE_WINDOW", 입력중지)

        def 큐알코드():
            교인이름, 교인, 교인뷰 = [], [], []
            명령 = "SELECT 이름, 교인코드 FROM " + DB테이블 + ";"
            cursor.execute(명령)
            for 값 in cursor.fetchall():
                교인이름.append([값["교인코드"], 값["이름"]])
            교인이름 = dict(교인이름)
            명령 = "SELECT 이름, 교인코드 FROM " + DB테이블 + ";"
            cursor.execute(명령)
            for 값 in cursor.fetchall():
                교인.append(값["교인코드"])
                교인뷰.append(교인이름[값["교인코드"]])
                for i in range(1, 7):
                    교인.append(값["교인코드"]+"-"+str(i))
                    교인뷰.append(교인이름[값["교인코드"]]+"-"+헌금형식[str(i)])
            for 뷰, txt in zip(교인뷰, 교인):
                큐알 = qr.make(txt)
                큐알.save("data/QRCord/"+뷰+".png")
            mb.showinfo("저장", "저장되었습니다.")

        def 출석체크():
            폰트 = tf.Font(family="맑은 고딕", size=15)
            def 입력중지():
                MsgBox = mb.askyesno('종료하시겠습니까?', '작성중에 종료한다면 저장되지 않습니다.', icon='warning')
                if MsgBox is True:
                    main.deiconify()
                    출석체크창.destroy()
            # noinspection PyUnusedLocal
            def 출석(self):
                if 출석받기.get() in 교인코드:
                    시트.append([이름[교인코드.index(출석받기.get())]+":"+출석받기.get(), 오늘날짜])
                    엑셀.save("data/" + 출입명부 + ".xlsx")
                    명령 = "INSERT INTO `" + 출석테이블 + "` (교인코드, 날짜) VALUES (%s, %s);"
                    cursor.execute(명령, (출석받기.get(), 오늘날짜))
                    church_db.commit()
                    출석받기.delete(0, 'end')
                    출석받기.focus()

            오늘날짜 = str(dt.datetime.now().date())
            출입명부 = 오늘날짜 + " 출입명부"
            try:
                엑셀 = xl.load_workbook("data/" + 출입명부 + ".xlsx", data_only=True)
            except:
                엑셀 = xl.Workbook()

            시트 = 엑셀.active
            시트.cell(1, 1, "이름:교인코드")
            시트.cell(1, 2, "날짜")

            명령 = "SELECT 이름, 교인코드 FROM " + DB테이블 + ";"
            리스트 = []
            교인코드 = []
            이름 = []
            cursor.execute(명령)
            for 값 in cursor.fetchall():
                리스트값 = ""
                for i in list(값.values()):
                    리스트값 += i
                이름.append(리스트값[:-22])
                교인코드.append(리스트값[-22:])
                리스트값 = 리스트값[:-22] + ":" + 리스트값[-22:]
                리스트.append(리스트값)
            
            main.withdraw()
            출석체크창 = tk.Toplevel(main)
            출석체크창.title("출석")
            출석체크창.geometry("285x35")
            출석체크창.resizable(False, False)
            출석체크창.iconbitmap('assets/church_icon.ico')

            출석받기 = tk.Entry(출석체크창, font=폰트, width=25)
            출석받기.pack()
            출석받기.bind("<Return>", 출석)
            출석받기.focus()

            출석체크창.protocol("WM_DELETE_WINDOW", 입력중지)

        main = tk.Tk()
        main.iconbitmap('assets/church_icon.ico')
        main.title("교회 관리")
        main.geometry("380x300")
        main.resizable(False, False)
        main_font = tf.Font(family="맑은 고딕", size=23)
        main_header = tk.Label(main, text="교회 관리", font=main_font)
        main_header.grid(row=0, column=0, columnspan=10)
        button_font = tf.Font(family="맑은 고딕", size=15)

        자동입력버튼 = tk.Button(main, text="자동으로 정보 입력", font=button_font, command=자동입력)
        자동입력버튼.grid(row=1, column=0)

        수동입력버튼 = tk.Button(main, text="수동으로 정보 입력", font=button_font, command=수동입력)
        수동입력버튼.grid(row=1, column=1)

        수정입력버튼 = tk.Button(main, text="정보 수정", font=button_font, command=수정입력)
        수정입력버튼.grid(row=2, column=0, sticky="w" + "e")

        헌금입력버튼 = tk.Button(main, text="헌금 입력", font=button_font, command=헌금입력)
        헌금입력버튼.grid(row=2, column=1, sticky="w" + "e")

        누계확인버튼 = tk.Button(main, text="누계 확인", font=button_font, command=누계확인)
        누계확인버튼.grid(row=3, column=0, sticky="w" + "e")

        메일보내기버튼 = tk.Button(main, text="메일 보내기", font=button_font, command=메일보내기)
        메일보내기버튼.grid(row=3, column=1, sticky="w" + "e")

        큐알코드버튼 = tk.Button(main, text="큐알코드", font=button_font, command=큐알코드)
        큐알코드버튼.grid(row=4, column=0, sticky="w" + "e")

        출석체크버튼 = tk.Button(main, text="출석체크", font=button_font, command=출석체크)
        출석체크버튼.grid(row=4, column=1, sticky="w" + "e")

        main.protocol("WM_DELETE_WINDOW", 종료)
        main.mainloop()

    Main()


church()
